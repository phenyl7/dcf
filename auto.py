import requests
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter

# Your API key
api_key = "cbqZ2QHLPY7B5zq0SVrBDxtLcri82HCQ"

# Base URL for FMP API
base_url = "https://financialmodelingprep.com/api/v3"

# Folder to save the Excel file
save_folder = "/Users/rr/fa"

# Function to fetch the cash flow statement
def fetch_cash_flow_statement(ticker):
    endpoint = f"{base_url}/cash-flow-statement/{ticker}?apikey={api_key}"

    response = requests.get(endpoint)
    if response.status_code == 200:
        data = response.json()
        df = pd.DataFrame(data)
        return df
    else:
        print(f"Failed to fetch cash flow data for {ticker}")
        return None

# Function to apply custom formatting for numbers
def apply_custom_format(ws_model):
    for row in ws_model.iter_rows(min_row=1, max_row=ws_model.max_row, min_col=1, max_col=ws_model.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'  # Format as thousands without decimals

def save_to_excel(ticker, cash_flow_df):
    file_name = os.path.join(save_folder, f"${ticker}.xlsx")
    wb = Workbook()
    ws_model = wb.active
    ws_model.title = "model"
    ws_main = wb.create_sheet(title="main")

    # Add rows to the "main" sheet
    ws_main['B2'] = "price"
    ws_main['B3'] = "shares"
    ws_main['B4'] = "mc"
    ws_main['B5'] = "cash"
    ws_main['B6'] = "debt"
    ws_main['B7'] = "ev"
    ws_main['B9'] = "net cash"
    ws_main['C9'] = "=C5-C6"

    # Specify names to keep
    names_to_keep = [
        "date",
        "calendarYear",
        "netIncome",
        "operatingCashFlow",
        "capitalExpenditure",
        "freeCashFlow",
    ]

    # Filter the DataFrame and drop the 'symbol' column if it exists
    filtered_df = cash_flow_df[names_to_keep].copy()

    # Convert numeric values to millions
    numeric_columns = ['netIncome', 'operatingCashFlow', 'capitalExpenditure', 'freeCashFlow']
    filtered_df[numeric_columns] = filtered_df[numeric_columns] / 1_000_000

    # Transpose the filtered DataFrame
    transposed_df = filtered_df.T

    # Write the transposed data starting at row 5
    for r_idx, (index, row) in enumerate(transposed_df.iterrows(), 5):
        # Write row name in column A
        ws_model.cell(row=r_idx, column=1, value=index)
        
        # Write row values starting from column B
        for c_idx, value in enumerate(row, 2):
            ws_model.cell(row=r_idx, column=c_idx, value=value)

    # Reverse the order of columns B through F while keeping column A intact
    max_row = ws_model.max_row
    for r_idx in range(5, max_row + 1):  # Start from row 5
        values = []
        # Collect values from columns B through F
        for c_idx in range(2, 7):  # B=2, F=6
            cell = ws_model.cell(row=r_idx, column=c_idx)
            values.append(cell.value)
        
        # Write back the reversed values
        for c_idx, value in enumerate(reversed(values), 2):
            ws_model.cell(row=r_idx, column=c_idx, value=value)

    # Set column widths
    ws_model.column_dimensions['A'].width = 30
    for col in ws_model.iter_cols(min_col=2, max_col=ws_model.max_column):
        ws_model.column_dimensions[col[0].column_letter].width = 10

    # Add header row content
    ws_model['C1'] = "fgr"
    ws_model['D1'] = "wacc"
    ws_model['E1'] = "tgr"
    ws_model['F1'] = "npv of fcf"
    ws_model['G1'] = "tv"
    ws_model['H1'] = "pv of tv"
    ws_model['I1'] = "ev"
    ws_model['J1'] = "net cash"
    ws_model['K1'] = "eq val"
    ws_model['L1'] = "shares"
    ws_model['M1'] = "implied $"
    ws_model['N1'] = "current $"
    ws_model['O1'] = "upside"
    ws_model['G5'] = "1"
    ws_model['H5'] = "=G5+1"
    ws_model['G6'] = "=F6+1"
    ws_model['D2'] = "15%"
    ws_model['E2'] = "2%"
    ws_model['I2'] = "=F2+H2"
    ws_model['J2'] = "=main!C9"
    ws_model['K2'] = "=I2+J2"
    ws_model['L2'] = "=main!C3"
    ws_model['M2'] = "=K2/L2"
    ws_model['N2'] = "=main!C2"
    ws_model['O2'] = "=M2/N2-1"
    ws_model['A12'] = "fcf yoy"
    ws_model['C12'] = "=(C10-B10)/ABS(B10)"
    ws_model['A14'] = "fcf avg gr"
    ws_model['A15'] = "fcf 4y avg gr"
    ws_model['A16'] = "last fcf gr"
    ws_model['A17'] = "conservative fcf gr"
    ws_model['A18'] = "very conservative fcf gr"
    ws_model['A21'] = "quarterly cash flow"
    ws_model['A22'] = "cffo"
    ws_model['A23'] = "capex"
    ws_model['A24'] = "fcf"
    ws_model['B21'] = "q1"
    ws_model['C21'] = "q2"
    ws_model['D21'] = "q3"
    ws_model['E21'] = "q4"
    ws_model['A28'] = "next FY projection"
    ws_model['B28'] = "sum of q fcf"
    ws_model['C28'] = "avg q fcf"
    ws_model['D28'] = "est FY fcf"
    ws_model['B29'] = "=SUM(24:24)"
    ws_model['C29'] = "=B29/N"
    ws_model['D29'] = "=C29*4"

    # Add formulas in G2 and H2
    ws_model['G2'] = "=(last fcf proj)*(1+tgr)/(wacc-tgr)"
    ws_model['H2'] = "=tv/(1+wacc)^n"

    # Apply custom number formatting for numbers
    apply_custom_format(ws_model)

    # Save the workbook
    wb.save(file_name)
    print(f"Cash flow statements for {ticker} saved to {file_name}")

# Main function to run the script
def main():
    ticker = input("Enter the ticker symbol: ").upper()
    cash_flow_df = fetch_cash_flow_statement(ticker)
    
    if cash_flow_df is not None:
        save_to_excel(ticker, cash_flow_df)
    else:
        print("No cash flow data found.")

if __name__ == "__main__":
    main()
